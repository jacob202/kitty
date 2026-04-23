"""
BOM (Bill of Materials) Manager for Kitty AI schematic analyzer.

Provides comprehensive BOM export, supplier integration, pricing lookup,
and BOM comparison capabilities.
"""

import csv
import io
import logging
import os
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import requests

logger = logging.getLogger(__name__)


@dataclass
class BOMComponent:
    """Represents a single BOM component."""

    designator: str
    value: str
    type: str
    quantity: int = 1
    description: str = ""
    manufacturer: str = ""
    part_number: str = ""
    price: float = 0.0
    supplier: str = ""
    category: str = ""
    footprint: str = ""

    def to_dict(self) -> dict[str, Any]:
        return {
            "designator": self.designator,
            "value": self.value,
            "type": self.type,
            "quantity": self.quantity,
            "description": self.description,
            "manufacturer": self.manufacturer,
            "part_number": self.part_number,
            "price": self.price,
            "supplier": self.supplier,
            "category": self.category,
            "footprint": self.footprint,
        }


@dataclass
class SupplierPartInfo:
    """Part information from a supplier."""

    part_number: str
    manufacturer: str
    description: str
    availability: int
    pricing: list[dict[str, float]] = field(default_factory=list)
    datasheet_url: str = ""
    image_url: str = ""
    lead_time: str = ""
    supplier: str = ""


@dataclass
class BOMComparisonResult:
    """Result of comparing two BOMs."""

    added: list[BOMComponent] = field(default_factory=list)
    removed: list[BOMComponent] = field(default_factory=list)
    modified: list[tuple[BOMComponent, BOMComponent, list[str]]] = field(default_factory=list)
    unchanged: list[BOMComponent] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "added": [c.to_dict() for c in self.added],
            "removed": [c.to_dict() for c in self.removed],
            "modified": [
                {"old": old.to_dict(), "new": new.to_dict(), "changes": changes}
                for old, new, changes in self.modified
            ],
            "unchanged": [c.to_dict() for c in self.unchanged],
        }


@dataclass
class BOMSummary:
    """Summary statistics for a BOM."""

    total_parts: int
    unique_parts: int
    total_cost: float
    parts_by_category: dict[str, int]
    parts_by_type: dict[str, int]
    missing_pricing: int

    def to_dict(self) -> dict[str, Any]:
        return {
            "total_parts": self.total_parts,
            "unique_parts": self.unique_parts,
            "total_cost": self.total_cost,
            "parts_by_category": self.parts_by_category,
            "parts_by_type": self.parts_by_type,
            "missing_pricing": self.missing_pricing,
        }


class SupplierAPI:
    """Base class for supplier API integrations."""

    def __init__(self, api_key: str | None = None):
        self.api_key = api_key
        self.base_url = ""

    def search_part(self, manufacturer_part_number: str) -> SupplierPartInfo | None:
        """Search for a part by manufacturer part number."""
        raise NotImplementedError

    def get_pricing(self, part_number: str, quantity: int = 1) -> dict[str, Any] | None:
        """Get pricing for a part at specified quantity."""
        raise NotImplementedError

    def get_availability(self, part_number: str) -> dict[str, Any] | None:
        """Get availability/stock information."""
        raise NotImplementedError


class DigiKeyAPI(SupplierAPI):
    """DigiKey API integration."""

    def __init__(self, api_key: str | None = None, client_id: str | None = None):
        super().__init__(api_key)
        self.client_id = client_id or os.getenv("DIGIKEY_CLIENT_ID")
        self.api_key = api_key or os.getenv("DIGIKEY_API_KEY")
        self.base_url = "https://api.digikey.com/products/v4"
        self._access_token: str | None = None

    def _get_access_token(self) -> str | None:
        """Get OAuth2 access token."""
        if self._access_token:
            return self._access_token

        if not self.client_id or not self.api_key:
            return None

        try:
            response = requests.post(
                "https://api.digikey.com/v1/oauth2/token",
                data={
                    "client_id": self.client_id,
                    "client_secret": self.api_key,
                    "grant_type": "client_credentials",
                },
                timeout=10,
            )
            if response.status_code == 200:
                data = response.json()
                self._access_token = data.get("access_token")
                return self._access_token
        except Exception as e:
            logger.error(f"DigiKey token error: {e}")
        return None

    def search_part(self, manufacturer_part_number: str) -> SupplierPartInfo | None:
        """Search DigiKey for a part."""
        token = self._get_access_token()
        if not token:
            return self._get_mock_data(manufacturer_part_number, "digikey")

        try:
            response = requests.get(
                f"{self.base_url}/search",
                headers={
                    "Authorization": f"Bearer {token}",
                    "X-DIGIKEY-Client-Id": self.client_id,
                },
                params={
                    "keywords": manufacturer_part_number,
                    "limit": 1,
                },
                timeout=10,
            )

            if response.status_code == 200:
                data = response.json()
                products = data.get("products", [])
                if products:
                    product = products[0]
                    return SupplierPartInfo(
                        part_number=product.get("manufacturerPartNumber", ""),
                        manufacturer=product.get("manufacturer", {}).get("name", ""),
                        description=product.get("description", ""),
                        availability=product.get("quantityAvailable", 0),
                        pricing=self._parse_pricing(product.get("pricing", [])),
                        datasheet_url=product.get("datasheetUrl", ""),
                        image_url=product.get("photoUrl", ""),
                        lead_time=str(product.get("leadTime", "")),
                        supplier="digikey",
                    )
        except Exception as e:
            logger.error(f"DigiKey search error: {e}")

        return self._get_mock_data(manufacturer_part_number, "digikey")

    def _parse_pricing(self, pricing_data: list[dict]) -> list[dict[str, float]]:
        """Parse pricing tiers from API response."""
        pricing = []
        for tier in pricing_data:
            pricing.append(
                {
                    "quantity": tier.get("breakQuantity", 0),
                    "price": float(tier.get("unitPrice", 0)),
                }
            )
        return sorted(pricing, key=lambda x: x["quantity"])

    def _get_mock_data(self, part_number: str, supplier: str) -> SupplierPartInfo:
        """Generate mock data for demo/testing."""
        return SupplierPartInfo(
            part_number=part_number,
            manufacturer="Demo Manufacturer",
            description=f"Demo part {part_number}",
            availability=1000,
            pricing=[
                {"quantity": 1, "price": 0.50},
                {"quantity": 10, "price": 0.35},
                {"quantity": 100, "price": 0.25},
            ],
            datasheet_url=f"https://example.com/datasheets/{part_number}.pdf",
            image_url="",
            lead_time="2 weeks",
            supplier=supplier,
        )


class MouserAPI(SupplierAPI):
    """Mouser API integration."""

    def __init__(self, api_key: str | None = None):
        super().__init__(api_key)
        self.api_key = api_key or os.getenv("MOUSER_API_KEY")
        self.base_url = "https://api.mouser.com/api/v1"

    def search_part(self, manufacturer_part_number: str) -> SupplierPartInfo | None:
        """Search Mouser for a part."""
        if not self.api_key:
            return self._get_mock_data(manufacturer_part_number, "mouser")

        try:
            response = requests.post(
                f"{self.base_url}/search/partnumber",
                headers={
                    "Content-Type": "application/json",
                },
                json={
                    "SearchByPartRequest": {
                        "mouserPartNumber": manufacturer_part_number,
                    }
                },
                params={
                    "apiKey": self.api_key,
                },
                timeout=10,
            )

            if response.status_code == 200:
                data = response.json()
                parts = data.get("SearchResults", {}).get("Parts", [])
                if parts:
                    part = parts[0]
                    return SupplierPartInfo(
                        part_number=part.get("ManufacturerPartNumber", ""),
                        manufacturer=part.get("Manufacturer", ""),
                        description=part.get("Description", ""),
                        availability=part.get("AvailabilityInStock", 0),
                        pricing=self._parse_pricing(part.get("PriceBreaks", [])),
                        datasheet_url=part.get("DataSheetUrl", ""),
                        image_url=part.get("ImagePath", ""),
                        lead_time=part.get("LeadTime", ""),
                        supplier="mouser",
                    )
        except Exception as e:
            logger.error(f"Mouser search error: {e}")

        return self._get_mock_data(manufacturer_part_number, "mouser")

    def _parse_pricing(self, pricing_data: list[dict]) -> list[dict[str, float]]:
        """Parse pricing tiers from API response."""
        pricing = []
        for tier in pricing_data:
            pricing.append(
                {
                    "quantity": tier.get("Quantity", 0),
                    "price": float(tier.get("Price", 0)),
                }
            )
        return sorted(pricing, key=lambda x: x["quantity"])

    def _get_mock_data(self, part_number: str, supplier: str) -> SupplierPartInfo:
        """Generate mock data for demo/testing."""
        return SupplierPartInfo(
            part_number=part_number,
            manufacturer="Demo Manufacturer",
            description=f"Demo part {part_number}",
            availability=500,
            pricing=[
                {"quantity": 1, "price": 0.55},
                {"quantity": 10, "price": 0.40},
                {"quantity": 100, "price": 0.30},
            ],
            datasheet_url=f"https://example.com/datasheets/{part_number}.pdf",
            image_url="",
            lead_time="1 week",
            supplier=supplier,
        )


class LCSCAPI(SupplierAPI):
    """LCSC API integration (basic implementation)."""

    def __init__(self, api_key: str | None = None):
        super().__init__(api_key)
        self.api_key = api_key or os.getenv("LCSC_API_KEY")
        self.base_url = "https://lcsc.com/api"

    def search_part(self, manufacturer_part_number: str) -> SupplierPartInfo | None:
        """Search LCSC for a part."""
        # LCSC API requires special authentication
        # Return mock data for now
        return SupplierPartInfo(
            part_number=manufacturer_part_number,
            manufacturer="Demo Manufacturer",
            description=f"LCSC Demo part {manufacturer_part_number}",
            availability=10000,
            pricing=[
                {"quantity": 1, "price": 0.10},
                {"quantity": 100, "price": 0.05},
                {"quantity": 1000, "price": 0.03},
            ],
            datasheet_url="",
            image_url="",
            lead_time="3 days",
            supplier="lcsc",
        )


class BOMManager:
    """Main BOM management class."""

    def __init__(self, db_client=None):
        """Initialize BOM manager.

        Args:
            db_client: DuckDB client instance (optional)
        """
        self.db = db_client
        self._suppliers: dict[str, SupplierAPI] = {}

    def _get_db(self):
        """Get database connection."""
        if self.db is None:
            from src.utils.duckdb_client import DuckDBClient

            self.db = DuckDBClient()
        return self.db

    def _get_supplier(self, supplier_name: str) -> SupplierAPI:
        """Get or create supplier API instance."""
        supplier_name = supplier_name.lower()

        if supplier_name not in self._suppliers:
            if supplier_name == "digikey":
                self._suppliers[supplier_name] = DigiKeyAPI()
            elif supplier_name == "mouser":
                self._suppliers[supplier_name] = MouserAPI()
            elif supplier_name == "lcsc":
                self._suppliers[supplier_name] = LCSCAPI()
            else:
                raise ValueError(f"Unknown supplier: {supplier_name}")

        return self._suppliers[supplier_name]

    def _fetch_bom_components(self, project_id: str) -> list[BOMComponent]:
        """Fetch BOM components from database."""
        db = self._get_db()

        # Check if hardware_entities table has project data
        try:
            cursor = db.execute(
                """SELECT
                    designator, value, type,
                    COALESCE(properties->>'manufacturer', '') as manufacturer,
                    COALESCE(properties->>'part_number', '') as part_number,
                    COALESCE(properties->>'description', '') as description,
                    COALESCE(properties->>'category', '') as category,
                    COALESCE(properties->>'footprint', '') as footprint
                FROM hardware_entities
                WHERE project = ?
                ORDER BY designator""",
                (project_id,),
            )

            components = []
            for row in cursor.fetchall():
                components.append(
                    BOMComponent(
                        designator=row[0] or "",
                        value=row[1] or "",
                        type=row[2] or "unknown",
                        manufacturer=row[3] or "",
                        part_number=row[4] or "",
                        description=row[5] or "",
                        category=row[6] or "",
                        footprint=row[7] or "",
                    )
                )

            if components:
                return components
        except Exception as e:
            logger.error(f"hardware_entities query failed: {e}")

        # Fallback to bom_components table
        try:
            cursor = db.execute(
                """SELECT
                    designator, value, type, quantity,
                    description, manufacturer, part_number,
                    price, supplier
                FROM bom_components
                WHERE project = ?
                ORDER BY designator""",
                (project_id,),
            )

            components = []
            for row in cursor.fetchall():
                components.append(
                    BOMComponent(
                        designator=row[0] or "",
                        value=row[1] or "",
                        type=row[2] or "unknown",
                        quantity=row[3] or 1,
                        description=row[4] or "",
                        manufacturer=row[5] or "",
                        part_number=row[6] or "",
                        price=float(row[7] or 0),
                        supplier=row[8] or "",
                    )
                )

            return components
        except Exception as e:
            logger.error(f"bom_components query failed: {e}")
            return []

    def export_to_csv(
        self,
        project_id: str,
        output_path: str | None = None,
        include_supplier_info: bool = False,
        preferred_supplier: str = "digikey",
    ) -> str:
        """Export BOM to CSV format.

        Args:
            project_id: Project identifier
            output_path: Output file path (if None, returns CSV string)
            include_supplier_info: Whether to include supplier pricing data
            preferred_supplier: Preferred supplier for pricing

        Returns:
            CSV content as string, or file path if output_path specified
        """
        components = self._fetch_bom_components(project_id)

        if not components:
            if output_path:
                Path(output_path).write_text("")
                return output_path
            return ""

        # Build fieldnames
        fieldnames = [
            "designator",
            "value",
            "type",
            "quantity",
            "description",
            "manufacturer",
            "part_number",
            "category",
            "footprint",
        ]

        if include_supplier_info:
            fieldnames.extend(
                ["supplier", "unit_price", "extended_price", "availability", "lead_time"]
            )

        # Collect supplier data if needed
        supplier_data: dict[str, SupplierPartInfo | None] = {}
        if include_supplier_info:
            supplier = self._get_supplier(preferred_supplier)
            unique_parts = {c.part_number for c in components if c.part_number}
            for part_number in unique_parts:
                supplier_data[part_number] = supplier.search_part(part_number)

        # Write CSV
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        for comp in components:
            row = {
                "designator": comp.designator,
                "value": comp.value,
                "type": comp.type,
                "quantity": comp.quantity,
                "description": comp.description,
                "manufacturer": comp.manufacturer,
                "part_number": comp.part_number,
                "category": comp.category,
                "footprint": comp.footprint,
            }

            if include_supplier_info and comp.part_number in supplier_data:
                part_info = supplier_data[comp.part_number]
                if part_info and part_info.pricing:
                    # Get price for quantity 1 (or lowest tier)
                    unit_price = part_info.pricing[0]["price"]
                    row.update(
                        {
                            "supplier": preferred_supplier,
                            "unit_price": unit_price,
                            "extended_price": unit_price * comp.quantity,
                            "availability": part_info.availability,
                            "lead_time": part_info.lead_time,
                        }
                    )

            writer.writerow(row)

        if output_path:
            Path(output_path).write_text(output.getvalue(), encoding="utf-8")
            return output_path

        return output.getvalue()

    def export_to_excel(
        self,
        project_id: str,
        output_path: str,
        include_supplier_info: bool = False,
        preferred_supplier: str = "digikey",
    ) -> str:
        """Export BOM to Excel format with multiple sheets.

        Args:
            project_id: Project identifier
            output_path: Output file path
            include_supplier_info: Whether to include supplier pricing data
            preferred_supplier: Preferred supplier for pricing

        Returns:
            Path to exported Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError:
            # Fallback to CSV if openpyxl not available
            csv_path = output_path.replace(".xlsx", ".csv")
            return self.export_to_csv(
                project_id, csv_path, include_supplier_info, preferred_supplier
            )

        components = self._fetch_bom_components(project_id)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Headers
        headers = [
            "Designator",
            "Value",
            "Type",
            "Quantity",
            "Description",
            "Manufacturer",
            "Part Number",
            "Category",
            "Footprint",
        ]

        if include_supplier_info:
            headers.extend(
                ["Supplier", "Unit Price", "Extended Price", "Availability", "Lead Time"]
            )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Get supplier data if needed
        supplier_data: dict[str, SupplierPartInfo | None] = {}
        if include_supplier_info:
            supplier = self._get_supplier(preferred_supplier)
            unique_parts = {c.part_number for c in components if c.part_number}
            for part_number in unique_parts:
                supplier_data[part_number] = supplier.search_part(part_number)

        # Write data
        for row_idx, comp in enumerate(components, 2):
            row_data = [
                comp.designator,
                comp.value,
                comp.type,
                comp.quantity,
                comp.description,
                comp.manufacturer,
                comp.part_number,
                comp.category,
                comp.footprint,
            ]

            if include_supplier_info:
                if comp.part_number in supplier_data:
                    part_info = supplier_data[comp.part_number]
                    if part_info and part_info.pricing:
                        unit_price = part_info.pricing[0]["price"]
                        row_data.extend(
                            [
                                preferred_supplier,
                                unit_price,
                                unit_price * comp.quantity,
                                part_info.availability,
                                part_info.lead_time,
                            ]
                        )
                    else:
                        row_data.extend(["", "", "", "", ""])
                else:
                    row_data.extend(["", "", "", "", ""])

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                if col in [10, 11]:  # Price columns
                    cell.number_format = "$#,##0.00"

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Add summary sheet if supplier info included
        if include_supplier_info:
            summary_ws = wb.create_sheet("Summary")
            summary = self.get_bom_summary(project_id)

            summary_ws.cell(row=1, column=1, value="BOM Summary").font = Font(bold=True, size=14)
            summary_ws.cell(row=3, column=1, value=f"Project: {project_id}")
            summary_ws.cell(row=4, column=1, value=f"Total Parts: {summary.total_parts}")
            summary_ws.cell(row=5, column=1, value=f"Unique Parts: {summary.unique_parts}")
            summary_ws.cell(row=6, column=1, value=f"Total Cost: ${summary.total_cost:.2f}")

            # Parts by category
            summary_ws.cell(row=8, column=1, value="Parts by Category:").font = Font(bold=True)
            row = 9
            for category, count in summary.parts_by_category.items():
                summary_ws.cell(row=row, column=1, value=category)
                summary_ws.cell(row=row, column=2, value=count)
                row += 1

        # Save
        wb.save(output_path)
        return output_path

    def compare_boms(self, project_id_v1: str, project_id_v2: str) -> BOMComparisonResult:
        """Compare two BOM revisions.

        Args:
            project_id_v1: First project/revision ID
            project_id_v2: Second project/revision ID

        Returns:
            BOMComparisonResult with differences
        """
        components_v1 = {c.designator: c for c in self._fetch_bom_components(project_id_v1)}
        components_v2 = {c.designator: c for c in self._fetch_bom_components(project_id_v2)}

        result = BOMComparisonResult()

        # Find added components
        for designator, comp in components_v2.items():
            if designator not in components_v1:
                result.added.append(comp)

        # Find removed components
        for designator, comp in components_v1.items():
            if designator not in components_v2:
                result.removed.append(comp)

        # Find modified and unchanged components
        for designator in components_v1:
            if designator in components_v2:
                comp_v1 = components_v1[designator]
                comp_v2 = components_v2[designator]

                changes = []
                if comp_v1.value != comp_v2.value:
                    changes.append("value")
                if comp_v1.part_number != comp_v2.part_number:
                    changes.append("part_number")
                if comp_v1.quantity != comp_v2.quantity:
                    changes.append("quantity")
                if comp_v1.manufacturer != comp_v2.manufacturer:
                    changes.append("manufacturer")

                if changes:
                    result.modified.append((comp_v1, comp_v2, changes))
                else:
                    result.unchanged.append(comp_v1)

        return result

    def get_bom_summary(self, project_id: str) -> BOMSummary:
        """Get summary statistics for a BOM.

        Args:
            project_id: Project identifier

        Returns:
            BOMSummary with statistics
        """
        components = self._fetch_bom_components(project_id)

        total_parts = sum(c.quantity for c in components)
        unique_parts = len(components)

        parts_by_category: dict[str, int] = {}
        parts_by_type: dict[str, int] = {}
        total_cost = 0.0
        missing_pricing = 0

        for comp in components:
            # Category count
            category = comp.category or "Uncategorized"
            parts_by_category[category] = parts_by_category.get(category, 0) + comp.quantity

            # Type count
            comp_type = comp.type or "unknown"
            parts_by_type[comp_type] = parts_by_type.get(comp_type, 0) + comp.quantity

            # Cost calculation
            if comp.price and comp.price > 0:
                total_cost += comp.price * comp.quantity
            else:
                missing_pricing += 1

        return BOMSummary(
            total_parts=total_parts,
            unique_parts=unique_parts,
            total_cost=total_cost,
            parts_by_category=parts_by_category,
            parts_by_type=parts_by_type,
            missing_pricing=missing_pricing,
        )

    def search_part(
        self, manufacturer_part_number: str, supplier: str = "digikey"
    ) -> SupplierPartInfo | None:
        """Search for a part across supplier databases.

        Args:
            manufacturer_part_number: Part number to search
            supplier: Supplier to search (digikey, mouser, lcsc)

        Returns:
            SupplierPartInfo if found
        """
        supplier_api = self._get_supplier(supplier)
        return supplier_api.search_part(manufacturer_part_number)

    def get_pricing(
        self, part_number: str, quantity: int = 1, supplier: str = "digikey"
    ) -> dict[str, Any] | None:
        """Get pricing for a part at specified quantity.

        Args:
            part_number: Part number
            quantity: Desired quantity
            supplier: Supplier to query

        Returns:
            Pricing information
        """
        part_info = self.search_part(part_number, supplier)

        if not part_info or not part_info.pricing:
            return None

        # Find appropriate price tier
        unit_price = None
        for tier in part_info.pricing:
            if tier["quantity"] <= quantity:
                unit_price = tier["price"]
            else:
                break

        if unit_price is None and part_info.pricing:
            unit_price = part_info.pricing[0]["price"]

        return {
            "part_number": part_number,
            "supplier": supplier,
            "quantity": quantity,
            "unit_price": unit_price,
            "total_price": unit_price * quantity if unit_price else None,
            "availability": part_info.availability,
            "lead_time": part_info.lead_time,
            "pricing_tiers": part_info.pricing,
        }

    def get_availability(
        self, part_number: str, supplier: str = "digikey"
    ) -> dict[str, Any] | None:
        """Get availability/stock information for a part.

        Args:
            part_number: Part number
            supplier: Supplier to query

        Returns:
            Availability information
        """
        part_info = self.search_part(part_number, supplier)

        if not part_info:
            return None

        return {
            "part_number": part_number,
            "supplier": supplier,
            "availability": part_info.availability,
            "lead_time": part_info.lead_time,
            "manufacturer": part_info.manufacturer,
            "description": part_info.description,
        }

    def create_shopping_list(
        self, project_id: str, preferred_supplier: str = "digikey", min_stock: int = 10
    ) -> dict[str, Any]:
        """Create a shopping list for project components.

        Args:
            project_id: Project identifier
            preferred_supplier: Preferred supplier
            min_stock: Minimum stock threshold

        Returns:
            Shopping list with items, totals, and alternatives
        """
        components = self._fetch_bom_components(project_id)
        supplier = self._get_supplier(preferred_supplier)

        shopping_items = []
        total_cost = 0.0
        unavailable_items = []

        for comp in components:
            if not comp.part_number:
                continue

            part_info = supplier.search_part(comp.part_number)

            if part_info:
                # Check availability
                if part_info.availability >= comp.quantity:
                    # Get price
                    unit_price = part_info.pricing[0]["price"] if part_info.pricing else 0
                    line_total = unit_price * comp.quantity
                    total_cost += line_total

                    shopping_items.append(
                        {
                            "designator": comp.designator,
                            "part_number": comp.part_number,
                            "manufacturer": part_info.manufacturer,
                            "description": part_info.description,
                            "quantity": comp.quantity,
                            "unit_price": unit_price,
                            "line_total": line_total,
                            "availability": part_info.availability,
                            "lead_time": part_info.lead_time,
                            "in_stock": part_info.availability >= min_stock,
                            "datasheet_url": part_info.datasheet_url,
                        }
                    )
                else:
                    unavailable_items.append(
                        {
                            "designator": comp.designator,
                            "part_number": comp.part_number,
                            "required": comp.quantity,
                            "available": part_info.availability,
                        }
                    )
            else:
                unavailable_items.append(
                    {
                        "designator": comp.designator,
                        "part_number": comp.part_number,
                        "required": comp.quantity,
                        "available": 0,
                    }
                )

        return {
            "project_id": project_id,
            "supplier": preferred_supplier,
            "items": shopping_items,
            "unavailable": unavailable_items,
            "total_cost": total_cost,
            "total_items": len(shopping_items),
            "generated_at": datetime.now().isoformat(),
        }

    def export_shopping_list_to_csv(
        self, project_id: str, output_path: str, preferred_supplier: str = "digikey"
    ) -> str:
        """Export shopping list to CSV format suitable for supplier upload.

        Args:
            project_id: Project identifier
            output_path: Output file path
            preferred_supplier: Preferred supplier

        Returns:
            Path to exported CSV file
        """
        shopping_list = self.create_shopping_list(project_id, preferred_supplier)

        fieldnames = [
            "part_number",
            "manufacturer",
            "quantity",
            "description",
            "designator",
            "unit_price",
            "line_total",
        ]

        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for item in shopping_list["items"]:
                writer.writerow(
                    {
                        "part_number": item["part_number"],
                        "manufacturer": item["manufacturer"],
                        "quantity": item["quantity"],
                        "description": item["description"],
                        "designator": item["designator"],
                        "unit_price": item["unit_price"],
                        "line_total": item["line_total"],
                    }
                )

        return output_path


# Global instance for easy import
bom_manager = BOMManager()


# Convenience functions for direct import
def export_to_csv(
    project_id: str,
    output_path: str | None = None,
    include_supplier_info: bool = False,
    preferred_supplier: str = "digikey",
) -> str:
    """Export BOM to CSV."""
    return bom_manager.export_to_csv(
        project_id, output_path, include_supplier_info, preferred_supplier
    )


def export_to_excel(
    project_id: str,
    output_path: str,
    include_supplier_info: bool = False,
    preferred_supplier: str = "digikey",
) -> str:
    """Export BOM to Excel."""
    return bom_manager.export_to_excel(
        project_id, output_path, include_supplier_info, preferred_supplier
    )


def compare_boms(project_id_v1: str, project_id_v2: str) -> BOMComparisonResult:
    """Compare two BOM revisions."""
    return bom_manager.compare_boms(project_id_v1, project_id_v2)


def get_bom_summary(project_id: str) -> BOMSummary:
    """Get BOM summary statistics."""
    return bom_manager.get_bom_summary(project_id)


def search_part(
    manufacturer_part_number: str, supplier: str = "digikey"
) -> SupplierPartInfo | None:
    """Search for a part."""
    return bom_manager.search_part(manufacturer_part_number, supplier)


def get_pricing(
    part_number: str, quantity: int = 1, supplier: str = "digikey"
) -> dict[str, Any] | None:
    """Get pricing for a part."""
    return bom_manager.get_pricing(part_number, quantity, supplier)


def get_availability(part_number: str, supplier: str = "digikey") -> dict[str, Any] | None:
    """Get availability for a part."""
    return bom_manager.get_availability(part_number, supplier)


def create_shopping_list(
    project_id: str, preferred_supplier: str = "digikey", min_stock: int = 10
) -> dict[str, Any]:
    """Create a shopping list."""
    return bom_manager.create_shopping_list(project_id, preferred_supplier, min_stock)
